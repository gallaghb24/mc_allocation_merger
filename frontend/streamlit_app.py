# Media Centre Allocation Merger
# Streamlit app to consolidate allocation exports only
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("❌ `openpyxl` is not installed. Run `pip install openpyxl`.")
    st.stop()

# ─────────── Constants ───────────
THIN_SIDE   = Side(style="thin", color="000000")
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE, bottom=THIN_SIDE)
ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
BOLD_FONT   = Font(bold=True)

KEY_COLS = [
    "Store Number", "Store Name", "Address Line 1", "Address Line 2", "City or Town",
    "County", "Country", "Post Code", "Region / Area", "Location Type", "Trading Format",
]

# Only the fields we can derive from the allocation files themselves
LABELS = [
    "Brief Description", "Total (inc Overs)", "Total Allocations", "Overs",
]
LABEL_COL_XL = KEY_COLS.index("Trading Format") + 1  # K (1-based)
ITEM_START_XL = LABEL_COL_XL + 1                     # L (1-based)

# ─────────── Helpers ───────────
def extract_alloc(file):
    """Read one allocation export – returns (DataFrame, meta dict)."""
    df = pd.read_excel(file, header=6, engine="openpyxl")

    # Normalise column headers – some exports include stray whitespace or
    # different capitalisation which would otherwise cause KeyErrors.
    df.columns = df.columns.str.strip()
    store_col = next((c for c in df.columns if c.strip().lower() == "store number"), None)
    if store_col is None:
        st.error("❌ Could not find a 'Store Number' column in the uploaded file.")
        st.stop()
    if store_col != "Store Number":
        df = df.rename(columns={store_col: "Store Number"})
    df["Store Number"] = pd.to_numeric(df["Store Number"], errors="coerce").astype("Int64")

    raw = pd.read_excel(file, header=None, engine="openpyxl")
    meta = {}
    for col in range(len(KEY_COLS), raw.shape[1]):
        ref = str(raw.iloc[6, col])
        if ref == "nan":
            continue
        meta[ref] = {
            "brief_description": raw.iloc[1, col],
            "overs": 0 if pd.isna(raw.iloc[4, col]) else raw.iloc[4, col],
        }
    return df, meta


def merge_allocations(dfs):
    if not dfs:
        return pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    num_cols = [c for c in combined.columns if c not in KEY_COLS]
    combined[num_cols] = combined[num_cols].apply(pd.to_numeric, errors="coerce")
    agg_rules = {c: ("first" if c in KEY_COLS else "sum")
                 for c in combined.columns if c != "Store Number"}
    return (combined
            .groupby("Store Number", as_index=False)
            .agg(agg_rules)
            .sort_values("Store Number")
            .reset_index(drop=True))


def build_workbook(df: pd.DataFrame, meta: dict, consolidated_on: str) -> BytesIO:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        STARTROW = len(LABELS) + 1          # pandas header → Excel row N
        df.to_excel(writer, index=False, sheet_name="Master Allocation", startrow=STARTROW)
        ws = writer.sheets["Master Allocation"]

        # Row 1 – Consolidation info
        ws.cell(row=1, column=1, value=consolidated_on).font = BOLD_FONT

        # Column widths & hide C–J
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
            if "C" <= col_letter <= "J":
                ws.column_dimensions[col_letter].hidden = True

        # Header rows (− now just four)
        item_cols = [c for c in df.columns if c not in KEY_COLS]
        for r_off, label in enumerate(LABELS):
            row_num = 2 + r_off
            lh = ws.cell(row=row_num, column=LABEL_COL_XL, value=label)
            lh.alignment = Alignment(vertical="center")
            lh.fill, lh.font, lh.border = ORANGE_FILL, BOLD_FONT, THIN_BORDER

            for idx, item in enumerate(item_cols):
                cell = ws.cell(row=row_num, column=ITEM_START_XL + idx)
                data  = meta.get(item, {})
                overs = data.get("overs", 0)
                total = df[item].fillna(0).sum()

                if label == "Brief Description":
                    cell.value = data.get("brief_description", "")
                elif label == "Total (inc Overs)":
                    cell.value = total + overs
                elif label == "Total Allocations":
                    cell.value = total
                elif label == "Overs":
                    cell.value = overs

                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

        # Style pandas header row
        header_excel_row = STARTROW + 1
        for col_idx in range(1, ws.max_column + 1):
            h = ws.cell(row=header_excel_row, column=col_idx)
            h.fill, h.font, h.border = ORANGE_FILL, BOLD_FONT, THIN_BORDER

        # Style data cells
        for row in ws.iter_rows(min_row=header_excel_row + 1,
                                max_row=ws.max_row,
                                min_col=1,
                                max_col=ws.max_column):
            for c in row:
                if c.column >= ITEM_START_XL:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = THIN_BORDER

    buffer.seek(0)
    return buffer

# ─────────── Streamlit UI ───────────
st.set_page_config(page_title="Media Centre Allocation Merger", layout="wide")
st.title("Media Centre Allocation Merger")

st.markdown(
    """**Step 1 – Download your allocation exports from Reports in Media Centre**  
**Step 2 – Drag and drop them below to upload**  
**Step 3 – Download the Consolidated Allocation**"""
)

alloc_files = st.file_uploader("Allocation exports (.xlsx)", type=["xlsx"], accept_multiple_files=True)

if not alloc_files:
    st.info("Please upload at least one allocation export.")
    st.stop()

ts = datetime.now()
consolidated_on = ts.strftime("Consolidated on %d/%m/%Y %H:%M")
file_ts = ts.strftime("%Y%m%d_%H%M")

# Merge process
progress  = st.progress(0)
all_dfs, meta = [], defaultdict(dict)
seen_refs, duplicates = set(), []
for idx, up in enumerate(alloc_files, start=1):
    df_part, meta_part = extract_alloc(up)
    item_cols = [c for c in df_part.columns if c not in KEY_COLS]
    new_cols = []
    for c in item_cols:
        if c in seen_refs:
            duplicates.append(c)
        else:
            seen_refs.add(c)
            new_cols.append(c)
    df_part = df_part[KEY_COLS + new_cols]
    meta_part = {k: v for k, v in meta_part.items() if k in new_cols}
    all_dfs.append(df_part)
    for k, v in meta_part.items():
        meta.setdefault(k, {}).update(v)
    progress.progress(idx / len(alloc_files))
progress.empty()

if duplicates:
    st.warning(
        "Duplicate brief reference(s) ignored: " + ", ".join(sorted(set(duplicates)))
    )

master_df = merge_allocations(all_dfs)
workbook  = build_workbook(master_df, meta, consolidated_on)

# Success summary & preview
lines_count = master_df.shape[1] - len(KEY_COLS)
st.success(f"Consolidated {lines_count} lines × {master_df.shape[0]} stores.")

# Download button
st.download_button(
    label="📥 Download Consolidated Allocation",
    data=workbook,
    file_name=f"Consolidated_Allocation_{file_ts}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
